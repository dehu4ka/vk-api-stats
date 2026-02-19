"""
Generate Excel report on archive integrity for all cameras.
Usage: python report.py
Output: report_YYYY-MM-DD_HHMMSS.xlsx
"""

import time
import sys
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from api.client import RTCameraClient
from services.stats import analyze_archive, format_duration

# --- config ---
PERIOD_DAYS = 7
WORKERS = 8
MAX_RETRIES = 3
RETRY_DELAY = 2  # seconds, doubles each retry
PROBLEM_COVERAGE_THRESHOLD = 50    # % — ниже = проблема
PROBLEM_MAX_GAP_THRESHOLD = 3600   # сек — выше = проблема
PROBLEM_DEPTH_THRESHOLD = 1        # дней — ниже = проблема

# --- styles ---
HEADER_FILL = PatternFill("solid", fgColor="2B3E50")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

GREEN_FILL = PatternFill("solid", fgColor="D4EDDA")
YELLOW_FILL = PatternFill("solid", fgColor="FFF3CD")
RED_FILL = PatternFill("solid", fgColor="F8D7DA")
GRAY_FILL = PatternFill("solid", fgColor="E2E3E5")

ONLINE_FONT = Font(color="0F5132", bold=True)
OFFLINE_FONT = Font(color="842029", bold=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="DEE2E6"),
    right=Side(style="thin", color="DEE2E6"),
    top=Side(style="thin", color="DEE2E6"),
    bottom=Side(style="thin", color="DEE2E6"),
)


def coverage_fill(pct):
    if pct >= 90:
        return GREEN_FILL
    elif pct >= 50:
        return YELLOW_FILL
    return RED_FILL


def style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def style_cell(cell):
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center")


def auto_width(ws, col_count, max_width=40):
    for col in range(1, col_count + 1):
        width = 0
        letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    width = max(width, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(width + 3, max_width)


_cancelled = threading.Event()


def fetch_fragments_with_retry(client, uid, since, till):
    for attempt in range(1, MAX_RETRIES + 1):
        if _cancelled.is_set():
            return []
        try:
            return client.get_camera_fragments(uid, since, till)
        except Exception as e:
            if attempt == MAX_RETRIES:
                raise
            delay = RETRY_DELAY * (2 ** (attempt - 1))
            # interruptible sleep
            if _cancelled.wait(delay):
                return []


def fetch_all_cameras_data(client):
    print("Fetching camera list...")
    cameras = client.get_all_cameras()
    total = len(cameras)
    print(f"  Found {total} cameras")
    print(f"  Workers: {WORKERS}, retries: {MAX_RETRIES}")

    now = time.time()
    now_int = int(now)
    since = now_int - PERIOD_DAYS * 86400

    results = [None] * total
    lock = threading.Lock()
    done = [0]
    errors = [0]

    def process_camera(idx, cam):
        if _cancelled.is_set():
            return
        uid = cam["uid"]
        name = cam.get("name") or uid[:12]

        try:
            fragments = fetch_fragments_with_retry(client, uid, since, now_int)
        except Exception as e:
            fragments = []
            with lock:
                errors[0] += 1
                sys.stdout.write(f"\n    ERROR {uid} ({name[:30]}): {e}\n")
                sys.stdout.flush()

        archive = analyze_archive(fragments, now)
        results[idx] = (cam, archive)

        with lock:
            done[0] += 1
            pct = done[0] / total * 100
            sys.stdout.write(
                f"\r  [{done[0]}/{total}] ({pct:.0f}%) "
                f"errors: {errors[0]}  "
                f"last: {name[:40]:<40}"
            )
            sys.stdout.flush()

    with ThreadPoolExecutor(max_workers=WORKERS) as pool:
        futures = {
            pool.submit(process_camera, i, cam): i
            for i, cam in enumerate(cameras)
        }
        try:
            for future in as_completed(futures):
                future.result()
        except KeyboardInterrupt:
            print("\n\n  Interrupted. Shutting down...")
            _cancelled.set()
            pool.shutdown(wait=True, cancel_futures=True)
            sys.exit(1)

    print(f"\n  Done. Processed {total} cameras, errors: {errors[0]}.")
    return results


def is_problem_camera(cam, archive):
    if archive["total_fragments"] == 0:
        return True
    if archive["coverage_pct"] < PROBLEM_COVERAGE_THRESHOLD:
        return True
    if archive["max_gap"] > PROBLEM_MAX_GAP_THRESHOLD:
        return True
    if archive["depth_days"] < PROBLEM_DEPTH_THRESHOLD:
        return True
    return False


def write_tldr_sheet(wb, data):
    ws = wb.active
    ws.title = "TLDR"

    TITLE_FONT = Font(bold=True, size=14)
    SECTION_FONT = Font(bold=True, size=11, color="2B3E50")
    LABEL_FONT = Font(color="555555")
    VALUE_FONT = Font(bold=True, size=12)
    TABLE_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    TABLE_HEADER_FILL = PatternFill("solid", fgColor="2B3E50")

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    total = len(data)
    online = sum(1 for c, _ in data if c.get("is_online"))
    offline = total - online
    with_archive = sum(1 for _, a in data if a["total_fragments"] > 0)
    no_archive = total - with_archive

    all_coverages = [a["coverage_pct"] for _, a in data if a["total_fragments"] > 0]
    avg_coverage = round(sum(all_coverages) / len(all_coverages), 1) if all_coverages else 0
    all_depths = [a["depth_days"] for _, a in data if a["total_fragments"] > 0]
    avg_depth = round(sum(all_depths) / len(all_depths), 1) if all_depths else 0

    green = sum(1 for c in all_coverages if c >= 90)
    yellow = sum(1 for c in all_coverages if 50 <= c < 90)
    red = sum(1 for c in all_coverages if 0 < c < 50)

    row = 2

    # --- title ---
    ws.cell(row=row, column=2, value="Archive Quality Report").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=2,
            value=f"Period: last {PERIOD_DAYS} days  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ).font = LABEL_FONT
    row += 2

    # --- overview ---
    ws.cell(row=row, column=2, value="Overview").font = SECTION_FONT
    row += 1

    labels_values = [
        ("Total cameras", total),
        ("Online / Offline", f"{online} / {offline}"),
        ("With archive / No archive", f"{with_archive} / {no_archive}"),
        ("Avg coverage", f"{avg_coverage}%"),
        ("Avg archive depth", f"{avg_depth} days"),
    ]
    for label, value in labels_values:
        ws.cell(row=row, column=2, value=label).font = LABEL_FONT
        c = ws.cell(row=row, column=3, value=value)
        c.font = VALUE_FONT
        row += 1

    row += 1

    # --- quality distribution ---
    ws.cell(row=row, column=2, value="Quality Distribution").font = SECTION_FONT
    row += 1

    dist = [
        ("Coverage >= 90%", green, GREEN_FILL),
        ("Coverage 50–89%", yellow, YELLOW_FILL),
        ("Coverage < 50%", red, RED_FILL),
        ("No archive", no_archive, GRAY_FILL),
    ]
    for label, count, fill in dist:
        ws.cell(row=row, column=2, value=label).font = LABEL_FONT
        c = ws.cell(row=row, column=3, value=count)
        c.font = VALUE_FONT
        c.fill = fill
        c.border = THIN_BORDER
        pct = round(count / total * 100, 1) if total else 0
        ws.cell(row=row, column=4, value=f"{pct}%").font = LABEL_FONT
        row += 1

    row += 1

    # --- worst coverage ---
    ws.cell(row=row, column=2, value="Worst Coverage (top 10)").font = SECTION_FONT
    row += 1

    worst_headers = ["Name", "Vendor", "DC", "Coverage %", "Max Gap", "Gaps"]
    for ci, h in enumerate(worst_headers, 2):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = TABLE_HEADER_FONT
        c.fill = TABLE_HEADER_FILL
        c.border = THIN_BORDER
    row += 1

    worst_cov = sorted(data, key=lambda x: x[1]["coverage_pct"])[:10]
    for cam, arc in worst_cov:
        cells = [
            cam.get("name") or cam["uid"][:12],
            cam.get("vendor") or "",
            (cam.get("data_center") or {}).get("name", ""),
            arc["coverage_pct"],
            format_duration(arc["max_gap"]) if arc["max_gap"] else "-",
            arc["gaps_count"],
        ]
        for ci, val in enumerate(cells, 2):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = THIN_BORDER
            if ci == 5:  # coverage
                c.fill = coverage_fill(arc["coverage_pct"])
        row += 1

    row += 1

    # --- most gaps ---
    ws.cell(row=row, column=2, value="Most Gaps (top 10)").font = SECTION_FONT
    row += 1

    gap_headers = ["Name", "Vendor", "DC", "Gaps", "Total Gap Time", "Coverage %"]
    for ci, h in enumerate(gap_headers, 2):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = TABLE_HEADER_FONT
        c.fill = TABLE_HEADER_FILL
        c.border = THIN_BORDER
    row += 1

    most_gaps = sorted(
        [(c, a) for c, a in data if a["gaps_count"] > 0],
        key=lambda x: x[1]["total_gap_time"],
        reverse=True,
    )[:10]
    for cam, arc in most_gaps:
        cells = [
            cam.get("name") or cam["uid"][:12],
            cam.get("vendor") or "",
            (cam.get("data_center") or {}).get("name", ""),
            arc["gaps_count"],
            format_duration(arc["total_gap_time"]) if arc["total_gap_time"] else "-",
            arc["coverage_pct"],
        ]
        for ci, val in enumerate(cells, 2):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = THIN_BORDER
        row += 1

    row += 1

    # --- longest single gap ---
    ws.cell(row=row, column=2, value="Longest Single Gap (top 10)").font = SECTION_FONT
    row += 1

    lgap_headers = ["Name", "Vendor", "DC", "Max Gap", "Gaps", "Coverage %"]
    for ci, h in enumerate(lgap_headers, 2):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = TABLE_HEADER_FONT
        c.fill = TABLE_HEADER_FILL
        c.border = THIN_BORDER
    row += 1

    longest_gap = sorted(
        [(c, a) for c, a in data if a["max_gap"] > 0],
        key=lambda x: x[1]["max_gap"],
        reverse=True,
    )[:10]
    for cam, arc in longest_gap:
        cells = [
            cam.get("name") or cam["uid"][:12],
            cam.get("vendor") or "",
            (cam.get("data_center") or {}).get("name", ""),
            format_duration(arc["max_gap"]),
            arc["gaps_count"],
            arc["coverage_pct"],
        ]
        for ci, val in enumerate(cells, 2):
            c = ws.cell(row=row, column=ci, value=val)
            c.border = THIN_BORDER
            if ci == 5:  # max gap
                if arc["max_gap"] > PROBLEM_MAX_GAP_THRESHOLD:
                    c.fill = RED_FILL
                elif arc["max_gap"] > 300:
                    c.fill = YELLOW_FILL
        row += 1


def write_summary_sheet(wb, data):
    ws = wb.create_sheet("Summary")

    headers = [
        "Name", "UID", "SN", "Vendor", "Model", "Address",
        "Data Center", "Status", "Depth (days)", "Recorded (h)",
        "Coverage %", "Fragments", "Avg Fragment", "Gaps > 1m",
        "Max Gap", "Total Gap Time",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    for cam, archive in data:
        status = "Online" if cam.get("is_online") else "Offline"
        row = [
            cam.get("name") or cam["uid"][:12],
            cam["uid"],
            cam.get("sn") or "",
            cam.get("vendor") or "",
            cam.get("model") or "",
            cam.get("address") or "",
            (cam.get("data_center") or {}).get("name", ""),
            status,
            archive["depth_days"],
            round(archive["total_recorded"] / 3600, 1) if archive["total_recorded"] else 0,
            archive["coverage_pct"],
            archive["total_fragments"],
            format_duration(archive["avg_fragment"]) if archive["avg_fragment"] else "-",
            archive["gaps_count"],
            format_duration(archive["max_gap"]) if archive["max_gap"] else "-",
            format_duration(archive["total_gap_time"]) if archive["total_gap_time"] else "-",
        ]
        ws.append(row)
        row_num = ws.max_row

        for col in range(1, len(headers) + 1):
            style_cell(ws.cell(row=row_num, column=col))

        # status coloring
        status_cell = ws.cell(row=row_num, column=8)
        if status == "Online":
            status_cell.font = ONLINE_FONT
            status_cell.fill = GREEN_FILL
        else:
            status_cell.font = OFFLINE_FONT
            status_cell.fill = RED_FILL

        # coverage coloring
        cov_cell = ws.cell(row=row_num, column=11)
        cov_cell.fill = coverage_fill(archive["coverage_pct"])

        # depth coloring
        depth_cell = ws.cell(row=row_num, column=9)
        if archive["depth_days"] < PROBLEM_DEPTH_THRESHOLD:
            depth_cell.fill = RED_FILL
        elif archive["depth_days"] < 3:
            depth_cell.fill = YELLOW_FILL

        # max gap coloring
        gap_cell = ws.cell(row=row_num, column=15)
        if archive["max_gap"] > PROBLEM_MAX_GAP_THRESHOLD:
            gap_cell.fill = RED_FILL
        elif archive["max_gap"] > 300:
            gap_cell.fill = YELLOW_FILL

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    auto_width(ws, len(headers))


def write_daily_sheet(wb, data):
    ws = wb.create_sheet("Daily")

    headers = [
        "Name", "UID", "Date", "Recorded (h)", "Coverage %",
        "Fragments", "Gaps > 1m", "Max Gap",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    for cam, archive in data:
        name = cam.get("name") or cam["uid"][:12]
        uid = cam["uid"]
        for day in archive["daily"]:
            row = [
                name,
                uid,
                day["date"],
                day["recorded_h"],
                day["coverage_pct"],
                day["fragments"],
                day["gaps_count"],
                format_duration(day["max_gap"]) if day["max_gap"] else "-",
            ]
            ws.append(row)
            row_num = ws.max_row

            for col in range(1, len(headers) + 1):
                style_cell(ws.cell(row=row_num, column=col))

            # coverage coloring
            cov_cell = ws.cell(row=row_num, column=5)
            cov_cell.fill = coverage_fill(day["coverage_pct"])

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    auto_width(ws, len(headers))


def write_problems_sheet(wb, data):
    ws = wb.create_sheet("Problems")

    headers = [
        "Name", "UID", "Vendor", "Model", "Address", "Data Center",
        "Status", "Depth (days)", "Coverage %", "Max Gap", "Reason",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    for cam, archive in data:
        if not is_problem_camera(cam, archive):
            continue

        reasons = []
        if archive["total_fragments"] == 0:
            reasons.append("No archive")
        if archive["coverage_pct"] < PROBLEM_COVERAGE_THRESHOLD:
            reasons.append(f"Low coverage ({archive['coverage_pct']}%)")
        if archive["max_gap"] > PROBLEM_MAX_GAP_THRESHOLD:
            reasons.append(f"Long gap ({format_duration(archive['max_gap'])})")
        if 0 < archive["depth_days"] < PROBLEM_DEPTH_THRESHOLD:
            reasons.append(f"Shallow depth ({archive['depth_days']}d)")

        status = "Online" if cam.get("is_online") else "Offline"
        row = [
            cam.get("name") or cam["uid"][:12],
            cam["uid"],
            cam.get("vendor") or "",
            cam.get("model") or "",
            cam.get("address") or "",
            (cam.get("data_center") or {}).get("name", ""),
            status,
            archive["depth_days"],
            archive["coverage_pct"],
            format_duration(archive["max_gap"]) if archive["max_gap"] else "-",
            "; ".join(reasons),
        ]
        ws.append(row)
        row_num = ws.max_row

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            style_cell(cell)

        # status coloring
        status_cell = ws.cell(row=row_num, column=7)
        if status == "Online":
            status_cell.font = ONLINE_FONT
            status_cell.fill = GREEN_FILL
        else:
            status_cell.font = OFFLINE_FONT
            status_cell.fill = RED_FILL

        # coverage
        ws.cell(row=row_num, column=9).fill = coverage_fill(archive["coverage_pct"])

        # reason highlight
        ws.cell(row=row_num, column=11).fill = RED_FILL

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    auto_width(ws, len(headers))


def main():
    client = RTCameraClient()

    print(f"=== RT Camera Archive Report ===")
    print(f"Period: last {PERIOD_DAYS} days")
    print()

    data = fetch_all_cameras_data(client)

    print("\nGenerating Excel...")
    wb = Workbook()
    write_tldr_sheet(wb, data)
    write_summary_sheet(wb, data)
    write_daily_sheet(wb, data)
    write_problems_sheet(wb, data)

    problem_count = sum(1 for cam, arc in data if is_problem_camera(cam, arc))

    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    filename = f"report_{timestamp}.xlsx"
    wb.save(filename)
    print(f"Saved: {filename}")
    print(f"  Total cameras:   {len(data)}")
    print(f"  Problem cameras: {problem_count}")


if __name__ == "__main__":
    main()
